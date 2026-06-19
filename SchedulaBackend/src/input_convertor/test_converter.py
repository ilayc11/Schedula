# src/input_convertor/test_converter.py
"""
Test converter — simulates the full pipeline and exports each DB table as an
Excel sheet with the real schema: simulated auto-increment PKs and fully
resolved FKs, so the output is identical to what would land in the database.

Sheets / DB tables:
  users            — user_internal_id*, user_id, user_name, first_name,
                     last_name, email, role, department_id
  courses          — course_id*, course_number, course_name,
                     department_id, degree_level, credit_points
  course_offering  — offering_id*, course_number, academic_year,
                     semester, group_number
  offering_cohorts — cohort_id*, offering_id (FK), target_department_id,
                     target_year_level
  lecturer_courses — lecturer_course_id*, lecturer_internal_id (FK),
                     offering_id (FK), role
  (* = simulated auto-increment PK)
"""

from typing import Dict, Any, List, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .excel_reader import read_lecturers_excel, read_courses_excel
from .mapper import map_lecturers, map_courses, group_cohorts_by_course


async def test_and_export_excel(
    lecturers_bytes: bytes,
    courses_bytes: bytes,
    output_path: str = "test_output.xlsx",
) -> Dict[str, Any]:
    """
    Run the full pipeline without API calls.
    Simulates DB auto-increment IDs and FK resolution, then exports each
    table as an Excel sheet so the result matches the real DB schema.
    """
    # ── 1. Read Excel files ──────────────────────────────────────────────────
    lecturers_df = await read_lecturers_excel(lecturers_bytes)
    courses_df = await read_courses_excel(courses_bytes)

    # ── 2. Map (same logic as production converter) ──────────────────────────
    users, offerings, lecturer_courses = map_lecturers(lecturers_df)
    courses, cohorts = map_courses(courses_df)

    # ── 3. Simulate DB auto-increment PKs & resolve FKs ─────────────────────

    # users → assign user_internal_id 1..N
    user_id_to_internal: Dict[str, int] = {}
    users_rows: List[Dict[str, Any]] = []
    for i, user in enumerate(users, start=1):
        user_id_to_internal[user["user_id"]] = i
        users_rows.append({
            "user_internal_id": i,
            "user_id": user["user_id"],
            "user_name": user["user_name"],
            "first_name": user["first_name"],
            "last_name": user["last_name"],
            "email": user["email"],
            "role": user["role"],
            "department_id": user["department_id"],
        })

    # courses → assign course_id 1..N
    courses_rows: List[Dict[str, Any]] = []
    for i, course in enumerate(courses, start=1):
        courses_rows.append({
            "course_id": i,
            "course_number": course["course_number"],
            "course_name": course["course_name"],
            "department_id": course["department_id"],
            "degree_level": course["degree_level"],
            "credit_points": course["credit_points"],
        })

    cohorts_by_course = group_cohorts_by_course(cohorts)

    # offerings → assign offering_id 1..N; build key→id map for FK resolution
    offering_key_to_id: Dict[Tuple[int, int, int, int], int] = {}
    offering_rows: List[Dict[str, Any]] = []
    for i, offering in enumerate(offerings, start=1):
        offering_rows.append({
            "offering_id": i,
            "course_number": offering["course_number"],
            "academic_year": offering["academic_year"],
            "semester": offering["semester"],
            "group_number": offering["group_number"],
        })
        key = (
            int(offering["course_number"]),
            int(offering["academic_year"]),
            int(offering["semester"]),
            int(offering["group_number"]),
        )
        offering_key_to_id[key] = i

    # offering_cohorts — one row per (offering × cohort of its course)
    cohort_rows: List[Dict[str, Any]] = []
    cohort_id = 1
    for offering in offering_rows:
        oid = offering["offering_id"]
        cn = int(offering["course_number"])
        for c in cohorts_by_course.get(cn, []):
            cohort_rows.append({
                "cohort_id": cohort_id,
                "offering_id": oid,
                "target_department_id": c["target_department_id"],
                "target_year_level": c["target_year_level"],
            })
            cohort_id += 1

    # lecturer_courses — resolve user_id → lecturer_internal_id
    #                    and (course,year,sem,group) → offering_id
    lc_rows: List[Dict[str, Any]] = []
    lc_id = 1
    for lc in lecturer_courses:
        lecturer_internal_id = user_id_to_internal.get(str(lc["user_id"]))
        key = (
            int(lc["course_number"]),
            int(lc["academic_year"]),
            int(lc["semester"]),
            int(lc["group_number"]),
        )
        offering_id = offering_key_to_id.get(key)
        if lecturer_internal_id is None or offering_id is None:
            continue  # unresolvable FK — same behaviour as production
        lc_rows.append({
            "lecturer_course_id": lc_id,
            "lecturer_internal_id": lecturer_internal_id,
            "offering_id": offering_id,
            "role": "Lecturer",
        })
        lc_id += 1

    # ── 4. Export to Excel ───────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    _add_sheet_from_df(wb, "users",            pd.DataFrame(users_rows))
    _add_sheet_from_df(wb, "courses",          pd.DataFrame(courses_rows))
    _add_sheet_from_df(wb, "course_offering",  pd.DataFrame(offering_rows))
    _add_sheet_from_df(wb, "offering_cohorts", pd.DataFrame(cohort_rows))
    _add_sheet_from_df(wb, "lecturer_courses", pd.DataFrame(lc_rows))

    wb.save(output_path)

    return {
        "status": "success",
        "message": f"Test data exported to {output_path}",
        "users_count": len(users_rows),
        "courses_count": len(courses_rows),
        "offerings_count": len(offering_rows),
        "offering_cohorts_count": len(cohort_rows),
        "lecturer_courses_count": len(lc_rows),
        "output_file": output_path,
    }


def _add_sheet_from_df(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)


if __name__ == "__main__":
    import asyncio
    with open("src/input_convertor/lecturer_full.xlsx", "rb") as f:
        lecturers_bytes = f.read()
    with open("src/input_convertor/courses_full.xlsx", "rb") as f:
        courses_bytes = f.read()
    result = asyncio.run(test_and_export_excel(lecturers_bytes, courses_bytes, "test_output_result_v2.xlsx"))
    print("=== TEST RESULT ===")
    for key, value in result.items():
        print(f"{key}: {value}")